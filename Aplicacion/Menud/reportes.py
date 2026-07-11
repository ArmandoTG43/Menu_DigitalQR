import base64
import calendar
import logging
from datetime import datetime, timedelta
from io import BytesIO
from collections import defaultdict
import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from django.db.models import Sum, Count, F
from django.http import FileResponse, HttpResponse
from django.shortcuts import render, redirect
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from .models import Pedido, DetallePedido, Producto
matplotlib.use('Agg')

logger = logging.getLogger(__name__)

# Constantes
DIAS_A_MOSTRAR = 7
COLOR_PRINCIPAL = '#e76e05'
COLOR_FONDO_GRAFICO = '#1f2937'
TOP_PRODUCTOS_LIMITE = 5
COLORES_CATEGORIAS = ['#e76e05', '#1f77b4', '#2ca02c', '#d62728', '#9467bd', '#8c564b']


def _generar_grafico_ventas(ventas_por_dia):
    if not ventas_por_dia:
        return None
    df = pd.DataFrame(list(ventas_por_dia))
    plt.figure(figsize=(10, 5))
    plt.plot(df['dia'], df['total'], marker='o', linewidth=2, color=COLOR_PRINCIPAL)
    plt.fill_between(df['dia'], df['total'], alpha=0.3, color=COLOR_PRINCIPAL)
    plt.title('Ventas de la última semana', fontsize=14, fontweight='bold')
    plt.xlabel('Fecha')
    plt.ylabel('Total Ventas ($)')
    plt.xticks(rotation=45)
    plt.grid(True, alpha=0.3)
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor=COLOR_FONDO_GRAFICO)
    buffer.seek(0)
    imagen_b64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    return imagen_b64


def _generar_grafico_productos(top_productos):
    if not top_productos:
        return None
    
    nombres = []
    cantidades = []
    for item in top_productos:
        nombre = item.get('producto__nombre') or item.get('nombre')
        if nombre:
            nombres.append(nombre)
            cantidades.append(item['total_vendido'])
    
    if not nombres:
        return None
    
    plt.figure(figsize=(10, 5))
    plt.barh(nombres, cantidades, color=COLOR_PRINCIPAL)
    plt.title('Productos más vendidos', fontsize=14, fontweight='bold')
    plt.xlabel('Cantidad vendida')
    plt.gca().invert_yaxis()
    
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor=COLOR_FONDO_GRAFICO)
    buffer.seek(0)
    imagen_b64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    return imagen_b64


def _generar_grafico_categorias(ventas_por_categoria):
    if not ventas_por_categoria:
        return None
    categorias = [item['producto__categoria'] for item in ventas_por_categoria]
    totales = [float(item['total']) for item in ventas_por_categoria]
    plt.figure(figsize=(7, 7))
    plt.pie(totales, labels=categorias, autopct='%1.1f%%',
            colors=COLORES_CATEGORIAS[:len(categorias)],
            startangle=90, shadow=True)
    plt.title('Ventas por categoría', fontsize=14, fontweight='bold')
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor=COLOR_FONDO_GRAFICO)
    buffer.seek(0)
    imagen_b64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    return imagen_b64


def _obtener_datos_semana(fecha_fin):
    dias = []
    ventas = []
    pedidos = []
    for i in range(DIAS_A_MOSTRAR - 1, -1, -1):
        fecha = fecha_fin - timedelta(days=i)
        nombre_dia = calendar.day_name[fecha.weekday()]
        dias.append(nombre_dia)
        total_dia = Pedido.objects.filter(
            fecha_hora__date=fecha.date(),
            estado='entregado'
        ).aggregate(total=Sum('total'))['total'] or 0
        ventas.append(float(total_dia))
        count_dia = Pedido.objects.filter(
            fecha_hora__date=fecha.date(),
            estado='entregado'
        ).count()
        pedidos.append(count_dia)
    return dias, ventas, pedidos
def _obtener_top_productos_con_imagenes(limite=TOP_PRODUCTOS_LIMITE):
    resultados = (
        DetallePedido.objects
        .filter(pedido__estado='entregado')
        .values('producto_id')
        .annotate(total_vendido=Sum('cantidad'))
        .order_by('-total_vendido')[:limite]
    )
    if not resultados:
        return []
    ids = [r['producto_id'] for r in resultados]
    cantidades = {r['producto_id']: r['total_vendido'] for r in resultados}
    productos = Producto.objects.filter(id__in=ids)
    top = []
    for producto in productos:
        imagen_url = None
        if hasattr(producto, 'imagen') and producto.imagen:
            try:
                imagen_url = producto.imagen.url
            except Exception:
                pass
        top.append({
            'producto__nombre': producto.nombre,  # 👈 CAMBIO AQUÍ
            'imagen_url': imagen_url,
            'total_vendido': cantidades.get(producto.id, 0),
        })
    top.sort(key=lambda x: x['total_vendido'], reverse=True)
    return top
def _obtener_ventas_por_categoria():
    if not hasattr(Producto, 'categoria'):
        return None
    if hasattr(Producto, 'precio'):
        return (
            DetallePedido.objects
            .filter(pedido__estado='entregado')
            .values('producto__categoria')
            .annotate(total=Sum(F('cantidad') * F('producto__precio')))
            .order_by('-total')
        )
    else:
        return (
            DetallePedido.objects
            .filter(pedido__estado='entregado')
            .values('producto__categoria')
            .annotate(total=Sum('cantidad'))
            .order_by('-total')
        )


def dashboard_ventas(request):
    if request.session.get('cliente'):
        return redirect('menu')

    fecha_fin = datetime.now()
    fecha_ini = fecha_fin - timedelta(days=DIAS_A_MOSTRAR)

    pedidos = Pedido.objects.filter(
        fecha_hora__range=[fecha_ini, fecha_fin],
        estado='entregado'
    )

    dias_semana, ventas_por_dia_chart, pedidos_por_dia_chart = _obtener_datos_semana(fecha_fin)

    ventas_por_dia = pedidos.extra(
        {'dia': "date(fecha_hora)"}
    ).values('dia').annotate(
        total=Sum('total'),
        cantidad=Count('id')
    ).order_by('dia')

    top_productos = _obtener_top_productos_con_imagenes()
    ventas_por_categoria = _obtener_ventas_por_categoria()

    grafico_ventas = _generar_grafico_ventas(ventas_por_dia)
    grafico_productos = _generar_grafico_productos(top_productos)
    grafico_categorias = _generar_grafico_categorias(ventas_por_categoria) if ventas_por_categoria else None

    total_ventas = pedidos.aggregate(Sum('total'))['total__sum'] or 0
    total_pedidos = pedidos.count()
    promedio = total_ventas / total_pedidos if total_pedidos > 0 else 0

    pedidos_pendientes = Pedido.objects.filter(estado='pendiente').count()
    pedidos_preparacion = Pedido.objects.filter(estado='en_preparacion').count()
    pedidos_listos = Pedido.objects.filter(estado='listo').count()

    productos_unicos = DetallePedido.objects.filter(
        pedido__estado='entregado',
        pedido__fecha_hora__range=[fecha_ini, fecha_fin]
    ).values('producto').distinct().count()

    context = {
        'grafico_ventas': grafico_ventas,
        'grafico_productos': grafico_productos,
        'grafico_categorias': grafico_categorias,
        'total_ventas': total_ventas,
        'total_pedidos': total_pedidos,
        'promedio': round(promedio, 2),
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_preparacion': pedidos_preparacion,
        'pedidos_listos': pedidos_listos,
        'top_productos': top_productos,
        'productos_unicos': productos_unicos,
        'dias': dias_semana,
        'ventas_por_dia_chart': ventas_por_dia_chart,
        'pedidos_por_dia_chart': pedidos_por_dia_chart,
    }
    return render(request, 'reportes/dashboard.html', context)

def reporte_ventas_excel(request):
    if request.session.get('cliente'):
        return redirect('menu')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws_ventas = wb.active
    ws_ventas.title = "Ventas"
    headers_ventas = ['ID', 'Mesa', 'Fecha', 'Total', 'Estado']
    for col, header in enumerate(headers_ventas, 1):
        celda = ws_ventas.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color=COLOR_PRINCIPAL.lstrip('#'),
                                 end_color=COLOR_PRINCIPAL.lstrip('#'),
                                 fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    pedidos = Pedido.objects.filter(estado='entregado').order_by('-fecha_hora')
    for row, pedido in enumerate(pedidos, 2):
        ws_ventas.cell(row=row, column=1, value=pedido.id)
        ws_ventas.cell(row=row, column=2, value=pedido.mesa.numero)
        ws_ventas.cell(row=row, column=3, value=pedido.fecha_hora.strftime('%Y-%m-%d %H:%M'))
        ws_ventas.cell(row=row, column=4, value=float(pedido.total))
        ws_ventas.cell(row=row, column=5, value=pedido.estado)
    for col in range(1, len(headers_ventas) + 1):
        ws_ventas.column_dimensions[get_column_letter(col)].width = 15

    ws_detalle = wb.create_sheet("Detalle")
    headers_detalle = ['Pedido ID', 'Mesa', 'Producto', 'Cantidad', 'Precio Unit.', 'Subtotal']
    for col, header in enumerate(headers_detalle, 1):
        celda = ws_detalle.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    detalles = DetallePedido.objects.filter(pedido__estado='entregado').select_related('producto', 'pedido', 'pedido__mesa')
    row = 2
    for det in detalles:
        precio = getattr(det.producto, 'precio', 0)
        ws_detalle.cell(row=row, column=1, value=det.pedido.id)
        ws_detalle.cell(row=row, column=2, value=det.pedido.mesa.numero)
        ws_detalle.cell(row=row, column=3, value=det.producto.nombre)
        ws_detalle.cell(row=row, column=4, value=det.cantidad)
        ws_detalle.cell(row=row, column=5, value=float(precio))
        ws_detalle.cell(row=row, column=6, value=float(precio * det.cantidad))
        row += 1
    for col in range(1, len(headers_detalle) + 1):
        ws_detalle.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas_detallado.xlsx"'
    wb.save(response)
    return response


def reporte_ventas_pdf(request):
    if request.session.get('cliente'):
        return redirect('menu')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elementos = []

    # Título y fecha
    titulo = Paragraph("REPORTE DE VENTAS AL DÍA", styles['Title'])
    fecha = Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elementos.append(titulo)
    elementos.append(fecha)
    elementos.append(Spacer(1, 0.3 * inch))

    # Obtener todos los pedidos entregados, ordenados por mesa y fecha
    pedidos = Pedido.objects.filter(estado='entregado').select_related('mesa').order_by('mesa__numero', 'fecha_hora')

    if not pedidos:
        elementos.append(Paragraph("No hay pedidos entregados para generar el reporte.", styles['Normal']))
        doc.build(elementos)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='reporte_ventas_por_mesa.pdf')

    # Agrupar por mesa
    mesas = {}
    for pedido in pedidos:
        mesa_num = pedido.mesa.numero
        if mesa_num not in mesas:
            mesas[mesa_num] = {
                'pedidos': [],
                'productos': defaultdict(lambda: {'cantidad': 0, 'precio': 0})
            }
        mesas[mesa_num]['pedidos'].append(pedido)

    # Procesar cada mesa: recorrer todos sus pedidos y acumular productos
    for mesa_num, data in mesas.items():
        for pedido in data['pedidos']:
            detalles = DetallePedido.objects.filter(pedido=pedido).select_related('producto')
            for det in detalles:
                producto = det.producto
                precio = getattr(producto, 'precio', 0)
                data['productos'][producto.nombre]['cantidad'] += det.cantidad
                # Guardamos el precio (asumimos que es el mismo para todas las ocurrencias)
                data['productos'][producto.nombre]['precio'] = precio

    # Ahora generar el PDF: una tabla por mesa
    for mesa_num, data in mesas.items():
        # Encabezado de la mesa
        # Obtener fechas de los pedidos de esta mesa para mostrarlas
        fechas = sorted([p.fecha_hora for p in data['pedidos']])
        fecha_str = fechas[0].strftime('%d/%m/%Y') if fechas else ''
        if len(fechas) > 1 and fechas[-1].date() != fechas[0].date():
            fecha_str += f" - {fechas[-1].strftime('%d/%m/%Y')}"
        header_text = f"Mesa {mesa_num} - {fecha_str}"
        elementos.append(Paragraph(header_text, styles['Heading2']))
        elementos.append(Spacer(1, 0.1 * inch))

        # Preparar datos de la tabla
        table_data = [['Producto', 'Cantidad', 'Precio Unit.', 'Subtotal']]
        total_mesa = 0
        for nombre, info in data['productos'].items():
            cant = info['cantidad']
            precio = info['precio']
            subtotal = cant * precio
            total_mesa += subtotal
            table_data.append([
                nombre,
                str(cant),
                f"${precio:.2f}",
                f"${subtotal:.2f}"
            ])
        # Fila de total de la mesa
        table_data.append(['', '', 'Total Mesa:', f"${total_mesa:.2f}"])

        # Crear tabla con anchos adecuados
        tabla = Table(table_data, colWidths=[3*inch, 0.8*inch, 1.2*inch, 1.2*inch])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -2), 1, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 0.3 * inch))

    doc.build(elementos)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='Reporte de Ventas.pdf')